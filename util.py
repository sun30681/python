#!/usr/bin/python
# -*- coding: UTF-8 -*-

import pandas as pd
import json
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension

def read_excel_to_json(file_path, sheet_name=0):
    """
    读取 Excel 文件并将其中的数据转换为 JSON 格式

    :param file_path: Excel 文件的路径
    :param sheet_name: 要读取的工作表名称或索引（默认为第一个工作表）
    :return: 包含 Excel 数据的 JSON 字符串
    """
    try:
        # 读取 Excel 文件
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 将 DataFrame 转换为 JSON 字符串
        json_data = df.to_json(orient='records', force_ascii=False, indent=4)

        return json_data
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return None

def json_to_excel(json_data, output_file, sheet_name='Sheet1'):
    """
    将 JSON 数据写入 Excel 文件并进行美化

    :param json_data: JSON 字符串或字典
    :param output_file: 输出 Excel 文件的路径
    :param sheet_name: 要写入的工作表名称（默认为 'Sheet1'）
    """
    try:
        # 如果 json_data 是字符串，先将其转换为字典
        if isinstance(json_data, str):
            data = json.loads(json_data)
        else:
            data = json_data

        # 将字典转换为 DataFrame
        df = pd.DataFrame(data)

        # 创建一个新的 Excel 工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 将 DataFrame 写入工作表
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 设置标题行样式
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment

        # 设置数据行样式
        data_font = Font(color="000000", name="Arial", size=11)
        data_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        data_alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.border = data_border
                cell.alignment = data_alignment

        # 自动调整列宽
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # 冻结标题行
        ws.freeze_panes = 'A2'

        # 设置页面边距
        ws.page_margins.left = 0.75
        ws.page_margins.right = 0.75
        ws.page_margins.top = 1.0
        ws.page_margins.bottom = 1.0

        # 保存工作簿
        wb.save(output_file)

        print(f"JSON data has been written to {output_file} with enhanced styling")
    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {e}")

# 示例用法
if __name__ == "__main__":
    # Excel 文件路径
    file_path = '运维服务部任务管理_运维自主承接情况.xlsx'

    # 读取 Excel 文件并转换为 JSON
    json_data = read_excel_to_json(file_path)

    if json_data is not None:
        print("Excel data as JSON:")
        print(json_data)

        # 输出 Excel 文件路径
        output_file = 'output.xlsx'

        # 将 JSON 数据写入 Excel 文件并进行美化
        json_to_excel(json_data, output_file)
